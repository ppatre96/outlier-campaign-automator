"""
Campaign Registry — tracks every campaign created by the pipeline in an Excel sheet.

Columns written at campaign creation time:
  smart_ramp_id, cohort_id, cohort_signature, geo_cluster, geo_cluster_label,
  geos, angle, campaign_type (inmail/static), advertised_rate,
  linkedin_campaign_urn, creative_urn, headline, subheadline, photo_subject,
  inmail_subject, created_at, status (active/deprecated/paused)

Metric columns updated later by the feedback agent:
  impressions, clicks, cpm_usd, ctr_pct, cpc_usd, spend_usd,
  applications, cpa_usd, last_metrics_at

The registry is written to data/campaign_registry.xlsx and also stored in
data/campaign_registry.json (for machine-readable access without openpyxl).

Pranav rule (2026-05-05): experimentation framework — max 3 cohorts per geo
cluster × 3 angles each. Feedback agent surfaces winners/losers; losing
angles are marked deprecated and replaced with new variants.
"""
from __future__ import annotations

import json
import logging
import threading
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

# Phase 3.3 — registry file I/O lock. Both the Static arm and the InMail
# arm (now run concurrently per Phase 3.3) call log_campaign / _load /
# _save against `data/campaign_registry.json` + `.xlsx`. The read-modify-
# write pattern (load → mutate → save) MUST be atomic across threads;
# without a lock, two arms can each load the same snapshot and overwrite
# each other's rows on save. The lock is re-entrant so a single thread
# can hold it across nested helper calls (e.g. log_campaign internally
# calling _save).
_registry_lock = threading.RLock()


@contextmanager
def registry_critical_section():
    """Acquire the registry lock for an atomic read-modify-write block.

    Use this when the caller needs to load → inspect → mutate → save the
    registry as one atomic operation (e.g. patching a creative URN onto
    a previously-logged row). Single-shot writes via log_campaign /
    update_metrics already acquire the lock internally; this context
    manager is for callers that need to hold the lock across multiple
    helper calls.
    """
    with _registry_lock:
        yield

_sheets_client = None

def _get_sheets():
    global _sheets_client
    if _sheets_client is None:
        try:
            from src.sheets import SheetsClient
            _sheets_client = SheetsClient()
        except Exception as exc:
            log.warning("Could not init SheetsClient for registry: %s", exc)
            from src.sheets import NullSheetsClient
            _sheets_client = NullSheetsClient()
    return _sheets_client

_REGISTRY_PATH = Path("data/campaign_registry.json")
_EXCEL_PATH    = Path("data/campaign_registry.xlsx")

COLUMNS = [
    # ── Identity ──────────────────────────────────────────────────────────────
    "smart_ramp_id",
    "cohort_id",
    "cohort_signature",
    "geo_cluster",
    "geo_cluster_label",
    "geos",
    "angle",                    # A / B / C / F
    "campaign_type",            # "static" | "inmail"
    "advertised_rate",          # e.g. "$50/hr"
    # ── Ad platform identity ──────────────────────────────────────────────────
    "channel",                  # display alias for platform — "LinkedIn" | "Meta" | "Google"
    "platform",                 # internal lower-case key — "linkedin" | "meta" | "google"
    "campaign_name",            # human-readable campaign name (matches platform UI exactly)
    "campaign_link",            # direct deep-link to the campaign in the platform's UI
    "platform_campaign_id",     # platform-native id (URN / numeric / resource name)
    "platform_creative_id",     # platform-native creative or ad id
    # ── LinkedIn URNs (legacy — duplicates platform_* for back-compat) ────────
    "linkedin_campaign_urn",
    "creative_urn",             # static: image creative URN; inmail: message ad URN
    # ── Copy snapshot ─────────────────────────────────────────────────────────
    "headline",
    "subheadline",
    "photo_subject",
    "creative_image_path",      # local path to the rendered PNG (used to embed image in Sheets)
    "cohort_geo",               # "<stg_id>__<geo_cluster>" — matches the Drive PNG folder name; used by reconcile_creative_paths() to backfill empty creative_image_path entries by walking Drive
    "inmail_subject",
    "inmail_body",              # full body delivered on LinkedIn
    # ── Lifecycle ─────────────────────────────────────────────────────────────
    "created_at",
    "status",                   # active | paused | deprecated
    "deprecation_reason",
    "gemini_prompt",            # base Gemini image gen prompt (first attempt, no QC suffix)
    # ── Performance metrics (filled by feedback agent) ─────────────────────────
    "impressions",
    "clicks",
    "cpm_usd",
    "ctr_pct",
    "cpc_usd",
    "spend_usd",
    "applications",
    "cpa_usd",
    "last_metrics_at",
    # ── 2026-05-20 additions (APPENDED at end on purpose) ─────────────────────
    # Inserting new columns into the middle of COLUMNS misreads every existing
    # data row by shifting positions. Always APPEND. The console parses by
    # header name (lib/sheets.ts) so position doesn't matter for new data.
    "audience_size",            # LinkedIn audience estimate at (cohort × geo) — populated by main.py per-geo audienceCounts recheck
    "qc_verdict",               # PASS | FAIL | "" — final QC verdict for this angle's creative
    "qc_attempts",              # int — gen attempts before final verdict
    "qc_violations",            # JSON list of violation strings — empty unless verdict=FAIL. Powers failure-analysis UI + per-rule skip overrides.
    # ── 2026-05-21 additions (APPEND ONLY) ───────────────────────────────────
    # Per-platform audience estimates from each channel's reach API. LinkedIn's
    # estimate stays in the existing `audience_size` column (back-compat); Meta
    # + Google get their own columns since estimate semantics differ per API.
    # Powers the per-channel AudienceBadge in the console.
    "meta_audience_size",
    "google_audience_size",
    "audience_check_status",    # "pass" | "denarrowed" | "below_floor" | "skipped" — outcome of audience_check.denarrow_for_platform
    # Funnel metrics joined nightly from Snowflake by feedback agent — null
    # until that join lands. Powers per-campaign cost-per-activation +
    # cost-per-worker-skill columns in the console's drilldown.
    "activations",              # count of contributors who completed onboarding for this campaign
    "skill_passes",             # count of contributors who passed at least one skill screen via this campaign
    # ── 2026-06-03 additions (APPEND ONLY) ───────────────────────────────────
    # Google search keywords attached to the Search ad-group for this row.
    # JSON-encoded list of strings (e.g. '["earn from home", "remote ai work", …]').
    # Populated by _process_extra_platform_arm when platform == "google".
    # Reviewed + editable in the outlier-campaign-console (keywords-card.tsx);
    # the console writes back to this column on save. Pipeline override-read on
    # next run is Phase 2 (TODO).
    "google_keywords",
    # ── 2026-07-06 addition (APPEND ONLY) ─────────────────────────────────────
    # Competitor-insight experiment id (e.g. "EXP-cardiologists-C") stamped on
    # the challenger-arm creative when a directive is pinned at creation time.
    # Lets the weekly readback attribute angle-C performance to the SPECIFIC
    # insight under test rather than to angle C in general. Empty on all
    # baseline (A/B) and non-experiment rows.
    "experiment_id",
]


@dataclass
class CampaignEntry:
    smart_ramp_id:          str = ""
    cohort_id:              str = ""
    cohort_signature:       str = ""
    geo_cluster:            str = ""
    geo_cluster_label:      str = ""
    geos:                   str = ""    # comma-joined ISO codes
    angle:                  str = ""
    campaign_type:          str = ""
    advertised_rate:        str = ""
    audience_size:          int | None = None  # LinkedIn audience count at (cohort × geo) — None when recheck failed or unavailable
    channel:                str = "LinkedIn"   # mirrors platform, title-cased for the Sheet view
    platform:               str = "linkedin"
    campaign_name:          str = ""    # human-readable campaign name (matches platform UI exactly)
    campaign_link:          str = ""    # deep-link to the campaign in the platform's UI
    platform_campaign_id:   str = ""
    platform_creative_id:   str = ""
    linkedin_campaign_urn:  str = ""    # legacy alias of platform_campaign_id when platform=linkedin
    creative_urn:           str = ""    # legacy alias of platform_creative_id
    headline:               str = ""
    subheadline:            str = ""
    photo_subject:          str = ""
    creative_image_path:    str = ""
    cohort_geo:             str = ""    # "<stg_id>__<geo_cluster>" — Drive PNG folder name
    inmail_subject:         str = ""
    inmail_body:            str = ""
    created_at:             str = ""
    status:                 str = "active"
    deprecation_reason:     str = ""
    gemini_prompt:          str = ""
    qc_verdict:             str = ""           # "PASS" | "FAIL" | "" when QC didn't run
    qc_attempts:            int | None = None  # gen attempts before final verdict
    qc_violations:          str = ""           # JSON-encoded list of violation strings
    meta_audience_size:     int | None = None  # Meta delivery_estimate midpoint at (cohort × geo)
    google_audience_size:   int | None = None  # Google audience-segment size sum at (cohort × geo)
    audience_check_status:  str = ""           # pass | denarrowed | below_floor | skipped
    activations:            int | None = None  # filled nightly from Snowflake
    skill_passes:           int | None = None  # filled nightly from Snowflake
    google_keywords:        str = ""           # JSON-encoded list of keyword strings; Google Search ad-group only
    # metrics — empty until feedback agent runs
    impressions:            int | None = None
    clicks:                 int | None = None
    cpm_usd:                float | None = None
    ctr_pct:                float | None = None
    cpc_usd:                float | None = None
    spend_usd:              float | None = None
    applications:           int | None = None
    cpa_usd:                float | None = None
    last_metrics_at:        str = ""
    experiment_id:          str = ""           # competitor-insight experiment id on challenger-arm rows


# Internal lower-case platform key → user-facing channel label shown in Sheet.
_CHANNEL_LABEL = {"linkedin": "LinkedIn", "meta": "Meta", "google": "Google"}


def _channel_label(platform: str) -> str:
    return _CHANNEL_LABEL.get((platform or "").lower(), (platform or "").title())


def _derive_campaign_link(platform: str, campaign_id: str) -> str:
    """Build the platform's UI deep-link from `platform_campaign_id`.

    LinkedIn: strips the URN prefix (`urn:li:sponsoredCampaign:` /
    `urn:li:sponsoredCampaignGroup:`) to the numeric id; campaign-group ids
    fall through the same campaigns/{id}/details URL — the Campaign Manager
    UI redirects to the correct entity type. Meta strips the `act_` prefix
    from META_AD_ACCOUNT_ID. Google needs no account id in the URL.
    """
    if not campaign_id:
        return ""
    p = (platform or "").lower()
    try:
        import config as _cfg
    except Exception:
        return ""
    if p == "linkedin":
        cid = str(campaign_id).rsplit(":", 1)[-1]
        return (
            f"https://www.linkedin.com/campaignmanager/accounts/"
            f"{_cfg.LINKEDIN_AD_ACCOUNT_ID}/campaigns/{cid}/details"
        )
    if p == "meta":
        meta_no_prefix = (_cfg.META_AD_ACCOUNT_ID or "").replace("act_", "")
        return (
            f"https://business.facebook.com/adsmanager/manage/campaigns"
            f"?act={meta_no_prefix}&selected_campaign_ids={campaign_id}"
        )
    if p == "google":
        return f"https://ads.google.com/aw/campaigns?campaignId={campaign_id}"
    return ""


def _load() -> list[dict]:
    # Lock acquisition is cheap on the RLock fast path; load is read-only
    # but we lock to prevent reading a half-written file mid-save.
    with _registry_lock:
        if _REGISTRY_PATH.exists():
            try:
                return json.loads(_REGISTRY_PATH.read_text())
            except Exception:
                pass
        return []


def _save(records: list[dict]) -> None:
    with _registry_lock:
        _REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
        _REGISTRY_PATH.write_text(json.dumps(records, indent=2, default=str))
        _write_excel(records)


def hydrate_from_postgres() -> int:
    """Overwrite the local JSON registry with the authoritative campaign rows
    from Postgres. Returns the number of rows hydrated.

    _load() reads a local file that is stale in CI (the committed
    data/campaign_registry.json lags the live pipeline). Postgres — written by
    upsert_campaign on every log_campaign — is always current. The metrics
    refresh calls this first so it iterates ALL current campaigns, not the
    stale committed subset. Best-effort: on read failure OR an empty result we
    keep the existing local registry rather than clobbering it with nothing."""
    try:
        from src.ui_decisions import list_all_campaign_data
        rows = list_all_campaign_data()
    except Exception as exc:
        log.warning("hydrate_from_postgres: read failed (%s) — keeping local registry", exc)
        return 0
    if not rows:
        log.warning("hydrate_from_postgres: Postgres returned 0 campaigns — keeping local registry")
        return 0
    _save(rows)
    log.info("hydrate_from_postgres: hydrated %d campaigns into local registry", len(rows))
    return len(rows)


def _write_excel(records: list[dict]) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.debug("openpyxl not installed — skipping Excel write (JSON only)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaign Registry"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Section dividers — colour-code by campaign_type
    static_fill  = PatternFill("solid", fgColor="E8F4FD")
    inmail_fill  = PatternFill("solid", fgColor="FEF9E7")
    depr_fill    = PatternFill("solid", fgColor="FDECEA")

    for row_idx, rec in enumerate(records, 2):
        fill = (depr_fill if rec.get("status") == "deprecated"
                else inmail_fill if rec.get("campaign_type") == "inmail"
                else static_fill)
        for col_idx, col in enumerate(COLUMNS, 1):
            val = rec.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False)

    # Column widths
    col_widths = {
        "smart_ramp_id": 14, "cohort_signature": 40, "geo_cluster_label": 22,
        "geos": 18, "angle": 7, "campaign_type": 10, "advertised_rate": 12,
        "linkedin_campaign_urn": 35, "creative_urn": 35,
        "headline": 30, "subheadline": 30, "photo_subject": 35,
        "inmail_subject": 35, "inmail_body": 80,
        "created_at": 20, "status": 12,
    }
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 14)

    ws.freeze_panes = "A2"
    _EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_EXCEL_PATH)
    log.info("Campaign registry written to %s (%d rows)", _EXCEL_PATH, len(records))


def log_campaign(
    smart_ramp_id: str,
    cohort_id: str,
    cohort_signature: str,
    geo_cluster: str,
    geo_cluster_label: str,
    geos: list[str],
    angle: str,
    campaign_type: str,
    advertised_rate: str,
    linkedin_campaign_urn: str = "",
    creative_urn: str = "",
    headline: str = "",
    subheadline: str = "",
    photo_subject: str = "",
    creative_image_path: str = "",
    inmail_subject: str = "",
    inmail_body: str = "",
    gemini_prompt: str = "",
    *,
    platform: str = "linkedin",
    platform_campaign_id: str = "",
    platform_creative_id: str = "",
    campaign_name: str = "",
    cohort_geo: str = "",
    audience_size: int | None = None,
    qc_verdict: str = "",
    qc_attempts: int | None = None,
    qc_violations: list[str] | None = None,
    meta_audience_size: int | None = None,
    google_audience_size: int | None = None,
    audience_check_status: str = "",
    google_keywords: list[str] | str | None = None,
    experiment_id: str = "",
) -> None:
    """Append one campaign row to the registry. Safe to call from any platform arm.

    `platform` defaults to "linkedin" so existing callers (which pass
    `linkedin_campaign_urn=` only) keep working. New multi-platform callers
    pass `platform=` plus `platform_campaign_id=` / `platform_creative_id=`.

    The legacy `linkedin_campaign_urn` / `creative_urn` columns are
    populated when platform="linkedin" so the existing readers + Google
    Sheets writers continue to work without further changes.
    """
    # Test-fixture leak guard (added 2026-05-27 after the audit caught 42
    # rows with `smart_ramp_id="flow"` + `urn:li:sponsoredCampaign:42` —
    # leaked from tests/test_inmail_isolation.py runs during the 13-hour
    # gap between e311ff1 and d511a28 before _block_registry_writes was
    # added). Reject any write carrying obvious test-mock values so future
    # test leaks fail loudly instead of silently polluting the registry.
    # If you legitimately need `smart_ramp_id="flow"` (e.g., a new dev tool)
    # rename your test fixture instead — flow is reserved for the leak path.
    _BANNED_RAMP_IDS = {"flow"}
    _BANNED_URNS = {
        "urn:li:sponsoredCampaign:42",
        "urn:li:sponsoredCreative:99",
    }
    if smart_ramp_id in _BANNED_RAMP_IDS:
        raise ValueError(
            f"log_campaign refusing to write: smart_ramp_id={smart_ramp_id!r} "
            f"is a known test-fixture value (see feedback_campaign_registry_singleton_leak "
            "in memory). If this is a real call, use a real GMR-XXXX id."
        )
    _suspect_urn = linkedin_campaign_urn or creative_urn or platform_campaign_id
    if _suspect_urn in _BANNED_URNS:
        raise ValueError(
            f"log_campaign refusing to write: URN {_suspect_urn!r} is a known "
            "test mock from tests/test_inmail_isolation.py. Likely cause: a "
            "test ran without `_block_registry_writes(monkeypatch)`. See memory "
            "feedback_campaign_registry_singleton_leak."
        )

    # Resolve the platform-native id pair from whichever set of kwargs the
    # caller used. LinkedIn callers may pass either set.
    pcid = platform_campaign_id or linkedin_campaign_urn
    pcrid = platform_creative_id or creative_urn

    # Central auto-stamp: challenger-arm creatives (angle C) carry the id of the
    # competitor insight that was pinned when they were generated. Stamping here
    # (the one choke point every platform arm calls) avoids threading the id
    # through dozens of call sites. Callers may still pass experiment_id
    # explicitly to override.
    if not experiment_id and (angle or "").upper() == "C":
        try:
            from src.competitor_experiment import active_directive
            _d = active_directive()
            if _d:
                experiment_id = _d.get("experiment_id", "") or ""
        except Exception:  # noqa: BLE001 — never block a registry write on this
            pass

    entry = CampaignEntry(
        smart_ramp_id=smart_ramp_id,
        cohort_id=cohort_id,
        cohort_signature=cohort_signature,
        geo_cluster=geo_cluster,
        geo_cluster_label=geo_cluster_label,
        geos=", ".join(geos) if geos else "",
        angle=angle,
        campaign_type=campaign_type,
        advertised_rate=advertised_rate,
        audience_size=audience_size,
        meta_audience_size=meta_audience_size,
        google_audience_size=google_audience_size,
        audience_check_status=audience_check_status,
        google_keywords=(
            json.dumps(list(google_keywords))
            if isinstance(google_keywords, list) and google_keywords
            else (google_keywords or "") if isinstance(google_keywords, str)
            else ""
        ),
        channel=_channel_label(platform),
        platform=platform,
        campaign_name=campaign_name,
        campaign_link=_derive_campaign_link(platform, pcid),
        platform_campaign_id=pcid,
        platform_creative_id=pcrid,
        # Legacy aliases — kept populated only for LinkedIn rows so old
        # consumers (Sheets columns, downstream queries) keep working.
        linkedin_campaign_urn=pcid if platform == "linkedin" else "",
        creative_urn=pcrid if platform == "linkedin" else "",
        headline=headline,
        subheadline=subheadline,
        photo_subject=photo_subject,
        creative_image_path=creative_image_path,
        cohort_geo=cohort_geo,
        inmail_subject=inmail_subject,
        inmail_body=inmail_body or "",
        gemini_prompt=gemini_prompt,
        qc_verdict=qc_verdict,
        qc_attempts=qc_attempts,
        qc_violations=json.dumps(qc_violations) if qc_violations else "",
        created_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        status="active",
        experiment_id=experiment_id,
    )
    # Hold the registry lock across the load-mutate-save window. _load and
    # _save also acquire the (re-entrant) lock internally; this ensures the
    # append is atomic w.r.t. concurrent writers from the other arm.
    with _registry_lock:
        records = _load()
        entry_dict = asdict(entry)
        records.append(entry_dict)
        _save(records)
    log.info(
        "Registry: logged %s/%s campaign %s (ramp=%s cohort=%s angle=%s geo=%s)",
        platform, campaign_type, pcid, smart_ramp_id, cohort_id, angle, geo_cluster_label,
    )
    try:
        _get_sheets().write_registry_row(entry_dict)
    except Exception as exc:
        log.warning("Registry sheet write failed (non-fatal): %s", exc)
    # Also persist to Postgres so the console can render Briefs & Campaigns
    # without the Sheet (which silently no-ops in CI when credentials.json is
    # absent). DATABASE_URL has no credentials.json dependency. Best-effort.
    try:
        from src.ui_decisions import upsert_campaign
        upsert_campaign(entry_dict)
    except Exception as exc:
        log.warning("Registry Postgres write failed (non-fatal): %s", exc)


def _id_match(rec: dict, campaign_id: str) -> bool:
    """True if the registry record matches `campaign_id` on either the new
    `platform_campaign_id` column or the legacy `linkedin_campaign_urn`."""
    return (
        rec.get("platform_campaign_id") == campaign_id
        or rec.get("linkedin_campaign_urn") == campaign_id
    )


def update_row(
    *,
    smart_ramp_id: str,
    cohort_geo: str,
    angle: str,
    platform: str,
    fields: dict,
) -> bool:
    """Patch a single registry row identified by
    (smart_ramp_id, cohort_geo, angle, platform). Used by the regen path to
    write back the new creative_image_path + creative_urn + qc_verdict +
    qc_violations after a successful re-gen.

    Returns True when exactly one row matched + was updated, False otherwise.
    Concurrent writers are safe (the file-level _registry_lock is held
    across load → mutate → save).
    """
    with _registry_lock:
        records = _load()
        updated = 0
        for rec in records:
            if (
                rec.get("smart_ramp_id") == smart_ramp_id
                and rec.get("cohort_geo") == cohort_geo
                and rec.get("angle") == angle
                and (rec.get("platform") or "").lower() == platform.lower()
            ):
                for k, v in fields.items():
                    rec[k] = v
                updated += 1
        if updated == 1:
            _save(records)
            log.info(
                "Registry: patched row ramp=%s cohort_geo=%s angle=%s platform=%s fields=%s",
                smart_ramp_id, cohort_geo, angle, platform, sorted(fields.keys()),
            )
            return True
        if updated == 0:
            log.warning(
                "Registry update_row found 0 matches for ramp=%s cohort_geo=%s angle=%s platform=%s",
                smart_ramp_id, cohort_geo, angle, platform,
            )
        else:
            log.warning(
                "Registry update_row matched %d rows for ramp=%s cohort_geo=%s angle=%s platform=%s — refusing to patch (ambiguous)",
                updated, smart_ramp_id, cohort_geo, angle, platform,
            )
        return False


def update_metrics(
    linkedin_campaign_urn: str,
    impressions: int,
    clicks: int,
    spend_usd: float,
    applications: int = 0,
) -> None:
    """Update performance metrics for a campaign. Called by the feedback agent.

    The first kwarg name is preserved for back-compat — it accepts any
    platform's campaign id (LinkedIn URN, Meta numeric, Google resource).
    """
    with _registry_lock:
        records = _load()
        matched = None
        for rec in records:
            if _id_match(rec, linkedin_campaign_urn):
                rec["impressions"] = impressions
                rec["clicks"] = clicks
                rec["spend_usd"] = round(spend_usd, 2)
                rec["cpm_usd"] = round(spend_usd / impressions * 1000, 2) if impressions else None
                rec["ctr_pct"] = round(clicks / impressions * 100, 3) if impressions else None
                rec["cpc_usd"] = round(spend_usd / clicks, 2) if clicks else None
                rec["applications"] = applications
                rec["cpa_usd"] = round(spend_usd / applications, 2) if applications else None
                rec["last_metrics_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
                matched = rec
                break
        if matched is not None:
            _save(records)
            log.info("Registry: metrics updated for %s", linkedin_campaign_urn)
        else:
            log.warning("Registry: campaign not found for metrics update: %s", linkedin_campaign_urn)
    # Also persist the updated metrics to Postgres so the console dashboard can
    # roll up impressions/spend/conversions without the Sheet (log_campaign
    # already dual-writes structure this way). Outside the registry lock —
    # mirrors log_campaign. Best-effort.
    if matched is not None:
        try:
            from src.ui_decisions import upsert_campaign
            upsert_campaign(matched)
        except Exception as exc:
            log.warning("Registry Postgres metrics write failed (non-fatal): %s", exc)


def update_funnel_metrics(
    creative_id: str,
    *,
    applications: int = 0,
    skill_passes: int = 0,
    activations: int = 0,
) -> int:
    """Write funnel outcomes (sign-ups / skill passes / activations) onto the
    registry row(s) for a LinkedIn creative, matched by platform creative id.

    These come from the FEED-15 funnel query (analyze_funnel_by_cohort), which
    attributes Outlier sign-ups → screening → activation to the individual
    creative via APPLICATION_CONVERSION.AD_ID + a linkedin/paid UTM filter. Ad
    reporting APIs only ever gave us impressions/clicks/spend; this is the leg
    that finally fills the previously-always-empty activation columns.

    `creative_id` accepts either a bare numeric id or a
    "urn:li:sponsoredCreative:<id>" URN. Recomputes cpa_usd from spend_usd when
    both are known. Dual-writes each matched row to Postgres so the console
    reflects it. Returns the number of rows updated."""
    cid = str(creative_id or "").rsplit(":", 1)[-1].strip()
    if not cid:
        return 0

    matched_rows: list[dict] = []
    with _registry_lock:
        records = _load()
        for rec in records:
            rec_cid = str(rec.get("platform_creative_id") or "").rsplit(":", 1)[-1].strip()
            if rec_cid and rec_cid == cid:
                rec["applications"] = applications
                rec["skill_passes"] = skill_passes
                rec["activations"] = activations
                spend = rec.get("spend_usd")
                if spend and applications:
                    rec["cpa_usd"] = round(float(spend) / applications, 2)
                rec["last_metrics_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
                matched_rows.append(rec)
        if matched_rows:
            _save(records)

    if not matched_rows:
        log.debug("Registry: no row for funnel creative_id=%s", cid)
        return 0

    # Mirror to Postgres (best-effort) so the console dashboard renders it.
    try:
        from src.ui_decisions import upsert_campaign
        for rec in matched_rows:
            upsert_campaign(rec)
    except Exception as exc:  # noqa: BLE001
        log.warning("Registry Postgres funnel write failed (non-fatal): %s", exc)

    log.info("Registry: funnel metrics on %d row(s) for creative_id=%s (apps=%d passes=%d act=%d)",
             len(matched_rows), cid, applications, skill_passes, activations)
    return len(matched_rows)


def reconcile_creative_paths(
    smart_ramp_id: str,
    platform: str = "linkedin",
    *,
    legacy_positional: bool = False,
    legacy_window_minutes: int = 60,
) -> dict:
    """Backfill empty `creative_image_path` entries by walking Drive.

    Pipeline writes registry rows at campaign-creation time. If the PNG render
    or Drive upload happens AFTER that (Gemini retry, async upload, slow Drive
    sync), the registry row ends up with an empty `creative_image_path` even
    though the PNG eventually lands in Drive at the canonical hierarchy
    `<ramp>/<platform>/<cohort_geo>/<angle>.png`.

    DEFAULT BEHAVIOR (safe — exact match only):
      For each row with `cohort_geo` set, look up the PNG at the row's
      `cohort_geo` + `angle` coordinates and patch the row. Rows without
      `cohort_geo` are left alone.

    OPT-IN LEGACY MODE (`legacy_positional=True`):
      For rows without `cohort_geo` (pre-cohort_geo-column rows), best-effort
      positional match by `(geo_cluster, angle)` ordered by `created_at`. Only
      assigns a PNG if its Drive `createdTime` is within `legacy_window_minutes`
      of the row's `created_at` — prevents pairing a 21:45 row to a 03:00 PNG
      just because they share a (geo, angle).

      Even with the window guard, legacy positional matching can mis-assign
      a PNG to the wrong row when multiple Smart Ramp rows mined the same
      (cohort × geo) pair (duplicate cohort matches with distinct campaigns).
      Use only when you accept that ambiguity.

    Returns a dict with counts of {patched, unmatched, ambiguous_legacy}.

    Idempotent — running twice patches nothing new on the second pass.
    """
    try:
        from src.gdrive import _service, _drive_id, _root_parent, find_or_create_folder
    except Exception as exc:  # pragma: no cover — Drive optional in tests
        log.warning("reconcile_creative_paths: Drive client unavailable: %s", exc)
        return {"patched": 0, "unmatched": 0, "ambiguous_legacy": 0}

    svc = _service()
    drive_id = _drive_id()

    def _list(parent_id: str) -> list[dict]:
        kw = {
            "q": f"'{parent_id}' in parents and trashed = false",
            "fields": "files(id,name,mimeType,webViewLink,createdTime)",
            "pageSize": 200,
            "supportsAllDrives": True,
            "includeItemsFromAllDrives": True,
        }
        if drive_id:
            kw["corpora"] = "drive"
            kw["driveId"] = drive_id
        return svc.files().list(**kw).execute().get("files", [])

    # Resolve Drive folder for this ramp+platform.
    root = _root_parent()
    try:
        ramp_folder = find_or_create_folder(smart_ramp_id, root, svc=svc)
        platform_folder = find_or_create_folder(platform, ramp_folder, svc=svc)
    except Exception as exc:
        log.warning(
            "reconcile_creative_paths(%s, %s): folder lookup failed: %s",
            smart_ramp_id, platform, exc,
        )
        return {"patched": 0, "unmatched": 0, "ambiguous_legacy": 0}

    # Walk: `<ramp>/<platform>/<cohort_geo>/<angle>.png`. Build:
    #   exact_by_key: {(cohort_geo, angle): drive_url}     — direct lookup
    #   legacy_by_geo_angle: {(geo_cluster, angle): [(created_at, url), ...]}
    #     used for positional fallback when cohort_geo is missing on the row.
    exact_by_key: dict[tuple[str, str], str] = {}
    legacy_by_geo_angle: dict[tuple[str, str], list[tuple[str, str]]] = {}

    for sub in _list(platform_folder):
        if sub.get("mimeType") != "application/vnd.google-apps.folder":
            continue
        sub_name = sub.get("name", "")
        # cohort_geo folder convention: "<stg_id>__<geo_cluster>"
        if "__" not in sub_name:
            continue
        geo_cluster = sub_name.rsplit("__", 1)[-1]
        for f in _list(sub["id"]):
            name = f.get("name", "")
            if not name.lower().endswith(".png"):
                continue
            angle = name.rsplit(".", 1)[0].strip()
            url = f.get("webViewLink", "")
            if not url:
                continue
            exact_by_key[(sub_name, angle)] = url
            legacy_by_geo_angle.setdefault((geo_cluster, angle), []).append(
                (f.get("createdTime", ""), url),
            )

    # Sort the legacy positional candidates by createdTime ascending.
    for k in legacy_by_geo_angle:
        legacy_by_geo_angle[k].sort(key=lambda t: t[0])

    patched = 0
    unmatched = 0
    ambiguous_legacy = 0

    with _registry_lock:
        records = _load()

        # Exact match by cohort_geo (new rows written post-this-fix). This is
        # always safe — folder name uniquely identifies the campaign row.
        for r in records:
            if r.get("smart_ramp_id") != smart_ramp_id:
                continue
            if (r.get("platform") or "linkedin") != platform:
                continue
            if r.get("creative_image_path"):
                continue
            cg = r.get("cohort_geo") or ""
            if not cg:
                continue  # legacy handled below (when opted in)
            url = exact_by_key.get((cg, r.get("angle", "")))
            if url:
                r["creative_image_path"] = url
                patched += 1
            else:
                unmatched += 1

        # Legacy positional fallback — explicit opt-in only. Pairs rows to
        # PNGs by (geo_cluster, angle) ordered by created_at, with a time-
        # window guard to prevent pairing a 21:45 row to a 03:00 PNG that
        # came from a different ramp run.
        if legacy_positional:
            from datetime import datetime as _dt, timedelta as _td

            def _row_ts(r):
                """Parse '2026-05-14 10:00 UTC' → datetime, or None."""
                ts = r.get("created_at", "") or ""
                try:
                    return _dt.strptime(ts, "%Y-%m-%d %H:%M UTC")
                except Exception:
                    return None

            def _png_ts(iso):
                try:
                    return _dt.strptime(iso[:19], "%Y-%m-%dT%H:%M:%S")
                except Exception:
                    return None

            # IDEMPOTENCY GUARD: collect URLs already assigned to ANY row in
            # the registry. We must not reassign them — otherwise running
            # reconcile twice produces duplicate URL ↔ row mappings. This
            # makes the legacy positional path safe to re-run.
            already_assigned: set[str] = {
                (r.get("creative_image_path") or "")
                for r in records
                if r.get("creative_image_path")
            }
            already_assigned.discard("")

            window = _td(minutes=legacy_window_minutes)
            legacy_rows = [
                r for r in records
                if r.get("smart_ramp_id") == smart_ramp_id
                and (r.get("platform") or "linkedin") == platform
                and not r.get("creative_image_path")
                and not r.get("cohort_geo")
            ]
            legacy_rows.sort(key=lambda r: r.get("created_at", ""))
            legacy_consumed: dict[tuple[str, str], set[int]] = {}
            for r in legacy_rows:
                key = (r.get("geo_cluster", ""), r.get("angle", ""))
                candidates = legacy_by_geo_angle.get(key, [])
                row_ts = _row_ts(r)
                consumed = legacy_consumed.setdefault(key, set())
                chosen_idx = None
                # Pick the EARLIEST not-yet-consumed candidate whose
                # createdTime is within `window` of the row's created_at
                # AND whose URL isn't already assigned elsewhere.
                for i, (iso, url) in enumerate(candidates):
                    if i in consumed or url in already_assigned:
                        continue
                    png_ts = _png_ts(iso)
                    if row_ts is None or png_ts is None:
                        chosen_idx = i
                        break
                    if abs(png_ts - row_ts) <= window:
                        chosen_idx = i
                        break
                if chosen_idx is not None:
                    chosen_url = candidates[chosen_idx][1]
                    r["creative_image_path"] = chosen_url
                    consumed.add(chosen_idx)
                    already_assigned.add(chosen_url)
                    patched += 1
                    if len(candidates) > 1:
                        ambiguous_legacy += 1
                else:
                    unmatched += 1

        if patched:
            _save(records)

    log.info(
        "reconcile_creative_paths(%s, %s): patched=%d unmatched=%d ambiguous_legacy=%d",
        smart_ramp_id, platform, patched, unmatched, ambiguous_legacy,
    )
    return {
        "patched":          patched,
        "unmatched":        unmatched,
        "ambiguous_legacy": ambiguous_legacy,
    }


def deprecate_campaign(linkedin_campaign_urn: str, reason: str) -> None:
    """Mark a campaign as deprecated. Accepts any platform's campaign id."""
    with _registry_lock:
        records = _load()
        for rec in records:
            if _id_match(rec, linkedin_campaign_urn):
                rec["status"] = "deprecated"
                rec["deprecation_reason"] = reason
                break
        _save(records)
    log.info("Registry: deprecated %s — %s", linkedin_campaign_urn, reason)


def get_active_campaigns(smart_ramp_id: str | None = None) -> list[dict]:
    """Return all active campaigns, optionally filtered by ramp."""
    records = _load()
    return [
        r for r in records
        if r.get("status") == "active"
        and (smart_ramp_id is None or r.get("smart_ramp_id") == smart_ramp_id)
    ]


def get_entry_by_urn(linkedin_campaign_urn: str) -> dict | None:
    """Return the registry entry for a campaign id, or None.

    Accepts either platform_campaign_id (new) or linkedin_campaign_urn (legacy).
    """
    records = _load()
    return next(
        (r for r in records if _id_match(r, linkedin_campaign_urn)),
        None,
    )


def get_cohort_entries(cohort_id: str, geo_cluster: str) -> list[dict]:
    """Return all registry entries for a cohort+geo combo, sorted by CTR desc."""
    records = _load()
    entries = [
        r for r in records
        if r.get("cohort_id") == cohort_id and r.get("geo_cluster") == geo_cluster
    ]
    return sorted(entries, key=lambda r: r.get("ctr_pct") or 0.0, reverse=True)


def get_registry_summary() -> dict:
    """Quick stats for Slack notifications / logging."""
    records = _load()
    return {
        "total": len(records),
        "active": sum(1 for r in records if r.get("status") == "active"),
        "deprecated": sum(1 for r in records if r.get("status") == "deprecated"),
        "by_ramp": {
            ramp: sum(1 for r in records if r.get("smart_ramp_id") == ramp)
            for ramp in {r.get("smart_ramp_id") for r in records}
        },
    }
